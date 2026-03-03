"""
Excelファイルから家計簿データをSQLiteにインポートするスクリプト

シート名がそのまま「対象月」（2602 → 2026年2月）。
シート内の取引日が別月の場合は「対象月の同日」として保存し（月末を超える場合は末日）、
備考に「実際: YYYY-MM-DD」を記録する。
"""
import sqlite3
import re
import calendar
from datetime import datetime, timedelta

import openpyxl

EXCEL_FILE = "20260301_ふたりの家計簿.xlsx"
DB = "household.db"

VALID_CATEGORIES = {
    "食費", "生活雑費", "家賃", "水道光熱費", "家具・家電", "その他",
    "外食", "デート", "旅行", "嗜好品",
}
VALID_PAYMENTS = {"共通カード", "あかり立替", "だいち立替"}


# シート名 → (year, month) のマッピング（シート名がそのまま対象月）
def parse_sheet_name(name):
    # 2603 → (2026, 3)
    m = re.fullmatch(r"(\d{2})(\d{2})", name)
    if m:
        return (2000 + int(m.group(1)), int(m.group(2)))

    # 202512 → (2025, 12)
    m = re.fullmatch(r"(20\d{2})(\d{2})", name)
    if m:
        return (int(m.group(1)), int(m.group(2)))

    # 2025年１月 → (2025, 1)
    m = re.search(r"(20\d{2})年\s*([０-９1-9１-９]+)月", name)
    if m:
        month_str = m.group(2)
        month = int(month_str.translate(str.maketrans("０１２３４５６７８９", "0123456789")))
        return (int(m.group(1)), month)

    # 11月, ９月, ... → need to guess the year (assume 2025)
    m = re.fullmatch(r"([０-９\d]+)月", name)
    if m:
        month_str = m.group(1)
        month = int(month_str.translate(str.maketrans("０１２３４５６７８９", "0123456789")))
        return (2025, month)

    return None


def resolve_amount(val):
    """数値セルまたは数式セル（=A+B-C形式）から金額を取得する"""
    if isinstance(val, (int, float)):
        return val
    if isinstance(val, str):
        val = val.strip()
        if val.startswith("="):
            expr = val[1:]
            # 四則演算と括弧のみ許可（安全な数式のみ評価）
            if re.fullmatch(r"[\d\+\-\*\/\.\(\)\s]+", expr):
                try:
                    return float(eval(expr))
                except Exception:
                    pass
    return None


def excel_serial_to_date(n):
    if isinstance(n, datetime):
        return n.strftime("%Y-%m-%d")
    if isinstance(n, (int, float)):
        dt = datetime(1899, 12, 30) + timedelta(days=int(n))
        return dt.strftime("%Y-%m-%d")
    if isinstance(n, str):
        n = n.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(n, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None


def init_db(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS expenses (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            amount INTEGER NOT NULL,
            category TEXT NOT NULL,
            detail TEXT,
            payment TEXT NOT NULL,
            notes TEXT,
            created_at TEXT DEFAULT (datetime('now', 'localtime'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS budgets (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            year INTEGER NOT NULL,
            month INTEGER NOT NULL,
            category TEXT NOT NULL,
            amount INTEGER NOT NULL,
            UNIQUE(year, month, category)
        )
    """)
    conn.commit()


SKIP_SHEETS = {"共通口座", "2024年11月・12月", "コピー用", "データ"}

# シート名が曖昧な場合の明示的な年月マッピング
# （parse_sheet_name では年を特定できないシートに使用）
SHEET_NAME_OVERRIDE = {
    "12月": (2024, 12),  # '202512' が 2025年12月なので、'12月' は 2024年12月
}

BUDGET_ROWS = {
    # row_idx (0-based) → subcategory
    3: "食費",
    4: "生活雑費",
    5: "家賃",
    6: "水道光熱費",
    7: "家具・家電",
    8: "その他",
    10: "外食",
    11: "デート",
    12: "旅行",
    13: "嗜好品",
}


def import_sheet(ws, year, month, conn):
    rows = list(ws.iter_rows(values_only=True))
    count = 0

    # Import budgets (rows 3-13, 0-based index)
    for row_idx, category in BUDGET_ROWS.items():
        if row_idx < len(rows):
            row = rows[row_idx]
            # Column D (index 3) = 予算
            budget_val = row[3] if len(row) > 3 else None
            if isinstance(budget_val, (int, float)) and budget_val > 0:
                conn.execute("""
                    INSERT OR REPLACE INTO budgets (year, month, category, amount)
                    VALUES (?, ?, ?, ?)
                """, (year, month, category, int(budget_val)))

    # Find data header row (contains '日付' or '金額' in column C)
    data_start = None
    for i, row in enumerate(rows):
        if len(row) > 2 and str(row[2]).strip() in ("金額", "日付"):
            data_start = i + 1
            break

    if data_start is None:
        return 0

    # Import transactions
    for row in rows[data_start:]:
        if not any(v is not None for v in row):
            continue

        # B=col1, C=col2, D=col3, E=col4, F=col5, G=col6
        date_raw = row[1] if len(row) > 1 else None
        amount_raw = row[2] if len(row) > 2 else None
        category = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        detail = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        payment = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        notes = str(row[6]).strip() if len(row) > 6 and row[6] else ""

        actual_date_str = excel_serial_to_date(date_raw)
        if not actual_date_str:
            continue
        amount_resolved = resolve_amount(amount_raw)
        if amount_resolved is None or amount_resolved <= 0:
            continue

        category = category.strip()
        if category not in VALID_CATEGORIES:
            continue
        if payment not in VALID_PAYMENTS:
            continue

        # 対象月と取引日の月が一致するかチェック
        actual_year = int(actual_date_str[:4])
        actual_month = int(actual_date_str[5:7])
        actual_day = int(actual_date_str[8:10])

        if actual_year == year and actual_month == month:
            # 同月 → 実際の日付をそのまま使用
            stored_date = actual_date_str
        else:
            # 別月 → 対象月の同日（月末超えはクランプ）として保存、備考に実際日付を記録
            last_day = calendar.monthrange(year, month)[1]
            stored_day = min(actual_day, last_day)
            stored_date = f"{year}-{month:02d}-{stored_day:02d}"
            actual_note = f"実際: {actual_date_str}"
            notes = f"{actual_note}, {notes}".strip(", ") if notes else actual_note

        conn.execute("""
            INSERT INTO expenses (date, amount, category, detail, payment, notes)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (stored_date, int(amount_resolved), category, detail, payment, notes))
        count += 1

    return count


def main():
    print(f"Excelファイルを読み込み中: {EXCEL_FILE}")
    # data_only=True でキャッシュ済み計算値を取得（数式セルも数値として読み込む）
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    conn = sqlite3.connect(DB)
    init_db(conn)

    # Clear existing data to avoid duplicates
    conn.execute("DELETE FROM expenses")
    conn.execute("DELETE FROM budgets")
    conn.commit()

    total = 0
    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        parsed = SHEET_NAME_OVERRIDE.get(sheet_name) or parse_sheet_name(sheet_name)
        if not parsed:
            print(f"  スキップ: {sheet_name} (年月不明)")
            continue

        year, month = parsed
        ws = wb[sheet_name]
        count = import_sheet(ws, year, month, conn)
        if count > 0:
            print(f"  {sheet_name} ({year}年{month}月): {count}件インポート")
            total += count

    conn.commit()
    conn.close()
    print(f"\n完了: 合計 {total} 件のデータをインポートしました")


if __name__ == "__main__":
    main()
